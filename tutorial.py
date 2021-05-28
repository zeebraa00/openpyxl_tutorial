import openpyxl

# load workbook
wb = openpyxl.load_workbook("./data.xlsx")
ws = wb["Sheet0"]

# make new workbook
wb2 = openpyxl.Workbook()
ws2 = wb2.create_sheet("new_sheet")

num=52745;i=1;cnt=0

while (True) :
    if (i == num) :
        break

    i+=1
    goal = 'AE'+str(i)
    chk = 'G'+str(i)
    if (ws[chk].value == "전문의약품") :
        cnt+=1
        continue

    try :
        curr = ws[goal].value.split("|")
    except :
        continue

    save = []

    leng = len(curr)
    for j in range(leng) :
        tstr = curr[j][9:]
        save.append(tstr)

    # write value to specific cell
    ws2.cell( row=i-cnt, column=1 ).value = save[0]

# save to xlsx file
wb2.save("./output.xlsx")
